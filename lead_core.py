"""
lead_core.py — Business logic for lead list processing.
All file I/O uses bytes / BytesIO so this module works in Pyodide (browser).
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

APP_VERSION = "0.1.0"

CANONICAL_COLUMNS = [
    "編號",
    "姓名",
    "ONE ID 序號",
    "LINEID",
    "手機(和泰會員)電銷",
    "手機(CR)電銷",
    "手機(和泰會員)SMS",
    "手機(CR)SMS",
    "訂單",
    "現保有車款_T",
    "現保有車交車年份_T",
    "現保有車款_L",
    "現保有車交車年份_L",
    "簡訊留名單",
    "LINE留名單",
]

OUTPUT_METADATA_COLUMNS = [
    "來源優先權",
    "來源群組",
    "來源檔案",
    "來源列號",
    "合併識別鍵",
    "更新時間",
]

MASTER_SHEET = "Working_List"
MERGED_SHEET = "Merged_Master"
DUPLICATE_SHEET = "Dropped_Duplicates"
FILTERED_DNC_SHEET = "Filtered_DNC"
RECENT_REMOVED_SHEET = "Removed_RecentOrders"
SHORT_URL_LOG_SHEET = "ShortURL_Log"
MANIFEST_SHEET = "Manifest"
README_SHEET = "README"

HEADER_ALIASES = {
    "編號": ["編號", "no", "序號"],
    "姓名": ["姓名", "name"],
    "ONE ID 序號": ["ONE ID 序號", "ONEID", "ONE ID", "ONE_ID", "ONE ID序號"],
    "LINEID": ["LINEID", "LINE ID", "LINE_ID"],
    "手機(和泰會員)電銷": ["手機(和泰會員)電銷", "手機(和泰會員) 電銷", "和泰會員電銷", "電話(和泰會員)電銷"],
    "手機(CR)電銷": ["手機(CR)電銷", "手機(CR) 電銷", "CR電銷", "電話(CR)電銷"],
    "手機(和泰會員)SMS": ["手機(和泰會員)SMS", "手機(和泰會員) SMS", "和泰會員SMS", "手機(和泰會員)簡訊"],
    "手機(CR)SMS": ["手機(CR)SMS", "手機(CR) SMS", "CRSMS", "手機(CR)簡訊", "手機(CR) sms"],
    "訂單": ["訂單", "MOBILE", "Phone", "PHONE", "手機", "行動電話"],
    "現保有車款_T": ["現保有車款_T", "現保有車款T"],
    "現保有車交車年份_T": ["現保有車交車年份_T", "現保有車交車年份T"],
    "現保有車款_L": ["現保有車款_L", "現保有車款L"],
    "現保有車交車年份_L": ["現保有車交車年份_L", "現保有車交車年份L"],
}

PHONE_FIELDS = [
    "手機(CR)SMS",
    "手機(和泰會員)SMS",
    "手機(CR)電銷",
    "手機(和泰會員)電銷",
]

PHONE_ELEC_FIELDS = ["手機(和泰會員)電銷", "手機(CR)電銷"]
PHONE_SMS_FIELDS = ["手機(和泰會員)SMS", "手機(CR)SMS"]

DNC_VALUES = {
    "不聯繫", "不連繫",
    "電話不聯繫", "電話不要打", "該電話不可聯繫",
    "簡訊不聯繫",
    "個資未授權",
}

NULL_LIKE = {
    "", "none", "null", "nan", "#n/a", "n/a",
    "電話不聯繫", "電話不要打", "該電話不可聯繫",
    "不聯繫", "不連繫",
    "簡訊不聯繫",
    "個資未授權",
    "無", "-", "--",
}

G_FILE_PATTERN = re.compile(r"(?:^|[^A-Z0-9])G(\d+)(?:_|\b|$)", re.IGNORECASE)
_LINE_ID_RE = re.compile(r"LINE\s*ID_?(U[A-Za-z0-9]+)", re.IGNORECASE)


@dataclass
class Record:
    data: Dict[str, str]
    source_file: str
    source_row: int
    priority: int
    group_name: str


class UnionFind:
    def __init__(self, n: int) -> None:
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x: int) -> int:
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a: int, b: int) -> None:
        ra = self.find(a)
        rb = self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            self.parent[ra] = rb
        elif self.rank[ra] > self.rank[rb]:
            self.parent[rb] = ra
        else:
            self.parent[rb] = ra
            self.rank[ra] += 1


# ---------------------------------------------------------------------------
# Text utilities
# ---------------------------------------------------------------------------

def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in NULL_LIKE:
        return ""
    return text


def normalize_header(text: str) -> str:
    text = clean_text(text)
    text = text.replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", "", text)
    return text.upper()


def normalize_phone(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    if len(digits) == 11 and digits.startswith("886"):
        digits = "0" + digits[-9:]
    elif len(digits) > 10 and digits.endswith(tuple(str(i) for i in range(10))):
        digits = digits[-10:]
    elif len(digits) == 9 and digits.startswith("9"):
        digits = "0" + digits
    if len(digits) < 8:
        return ""
    return digits


def normalize_id(value) -> str:
    return clean_text(value).upper()


def file_sha1_bytes(content: bytes) -> str:
    return hashlib.sha1(content).hexdigest()


def extract_priority(filename: str) -> Optional[int]:
    match = G_FILE_PATTERN.search(filename)
    if not match:
        return None
    return int(match.group(1))


def extract_group_name(filename: str) -> str:
    p = extract_priority(filename)
    return f"G{p}" if p is not None else ""


# ---------------------------------------------------------------------------
# Excel / CSV helpers
# ---------------------------------------------------------------------------

def map_headers(headers: List[str]) -> Dict[str, int]:
    normalized = [normalize_header(h) for h in headers]
    result: Dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        alias_set = {normalize_header(a) for a in aliases}
        for idx, header in enumerate(normalized):
            if header in alias_set:
                result[canonical] = idx
                break
    return result


def _is_dnc(raw_row: List[str], header_map: Dict[str, int], fields: List[str]) -> bool:
    for f in fields:
        if f not in header_map:
            continue
        idx = header_map[f]
        if idx < len(raw_row) and raw_row[idx].strip() in DNC_VALUES:
            return True
    return False


def read_excel_records_bytes(
    filename: str,
    file_bytes: bytes,
    filter_mode: str = "none",
) -> Tuple[List[Record], List[Record]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    all_rows = list(ws.iter_rows(values_only=True))
    # Find header row: first row with ≥2 non-empty cells
    header_idx = 0
    header_row_values: List[str] = []
    for i, row in enumerate(all_rows):
        values = [clean_text(cell) for cell in row]
        if sum(1 for v in values if v) >= 2:
            header_idx = i
            header_row_values = values
            break

    header_map = map_headers(header_row_values)
    if not header_map:
        return [], []

    priority = extract_priority(filename) or 9999
    group_name = extract_group_name(filename) or filename

    results: List[Record] = []
    filtered: List[Record] = []

    for row_idx, row in enumerate(all_rows):
        if row_idx <= header_idx:
            continue
        raw_row = [str(cell).strip() if cell is not None else "" for cell in row]
        row_values = [clean_text(cell) for cell in row]
        if not any(row_values):
            continue
        record_data = {col: "" for col in CANONICAL_COLUMNS}
        for canonical, col_idx in header_map.items():
            if col_idx < len(row_values):
                value = clean_text(row_values[col_idx])
                if canonical in PHONE_FIELDS:
                    value = normalize_phone(value)
                record_data[canonical] = value
        if not any(record_data.values()):
            continue
        phones_empty = not any(record_data.get(f, "") for f in PHONE_FIELDS)
        line_empty = not record_data.get("LINEID", "")
        if phones_empty and line_empty:
            continue
        rec = Record(
            data=record_data,
            source_file=filename,
            source_row=row_idx + 1,
            priority=priority,
            group_name=group_name,
        )
        if filter_mode != "none":
            phone_dnc = _is_dnc(raw_row, header_map, PHONE_ELEC_FIELDS)
            sms_dnc = _is_dnc(raw_row, header_map, PHONE_SMS_FIELDS)
            should_filter = (
                (filter_mode == "phone" and phone_dnc)
                or (filter_mode == "sms" and sms_dnc)
                or (filter_mode == "either" and (phone_dnc or sms_dnc))
            )
            if should_filter:
                filtered.append(rec)
                continue
        results.append(rec)
    return results, filtered


def read_csv_rows_bytes(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    """Try multiple encodings and return (headers, rows)."""
    for enc in ["utf-8-sig", "utf-8", "cp950", "big5"]:
        try:
            text = content.decode(enc)
            reader = csv.DictReader(io.StringIO(text, newline=""))
            headers = list(reader.fieldnames or [])
            rows = [{k or "": clean_text(v) for k, v in row.items()} for row in reader]
            return headers, rows
        except Exception:
            pass
    raise RuntimeError("無法解析 CSV 檔案，請確認編碼格式。")


# ---------------------------------------------------------------------------
# Workbook utilities
# ---------------------------------------------------------------------------

def _excel_col(index: int) -> str:
    result = ""
    while index:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


def auto_width(ws) -> None:
    widths: Dict[int, int] = defaultdict(lambda: 10)
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            text = clean_text(value)
            if not text:
                continue
            widths[idx] = min(max(widths[idx], len(text) + 2), 40)
    for idx, width in widths.items():
        ws.column_dimensions[_excel_col(idx)].width = width


def recreate_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(title=name)


def sheet_to_dicts(ws) -> List[Dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_text(h) for h in rows[0]]
    results = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else ""
            item[header] = clean_text(value)
        if any(item.values()):
            results.append(item)
    return results


def write_dicts_sheet(wb: Workbook, name: str, rows: List[Dict[str, str]], headers: List[str]) -> None:
    ws = recreate_sheet(wb, name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    auto_width(ws)
    ws.freeze_panes = "A2"


def write_readme_sheet(wb: Workbook) -> None:
    ws = recreate_sheet(wb, README_SHEET)
    rows = [
        ["本工具用途", "本機端整理名單，無須上傳外部服務"],
        ["版本", APP_VERSION],
        ["Working_List", "目前對外使用中的名單"],
        ["Merged_Master", "原始合併結果備份"],
        ["Dropped_Duplicates", "因優先權或重複被捨棄的名單"],
        ["Filtered_DNC", "因電話/SMS 標示不聯繫而被過濾的名單"],
        ["Removed_RecentOrders", "因近年受訂而被排除的名單"],
        ["ShortURL_Log", "短網址配對紀錄"],
        ["Manifest", "本次使用到的來源檔案與雜湊"],
    ]
    for row in rows:
        ws.append(row)
    auto_width(ws)


def write_manifest_sheet(wb: Workbook, manifest_rows: List[Dict[str, str]]) -> None:
    ws = recreate_sheet(wb, MANIFEST_SHEET)
    headers = ["operation", "file_name", "size", "sha1", "processed_time"]
    ws.append(headers)
    for row in manifest_rows:
        ws.append([row.get(h, "") for h in headers])
    auto_width(ws)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load_wb(wb_bytes: bytes) -> Workbook:
    return load_workbook(io.BytesIO(wb_bytes))


# ---------------------------------------------------------------------------
# Deduplication helpers
# ---------------------------------------------------------------------------

def record_identifiers(record: Record) -> List[str]:
    keys: List[str] = []
    for field in PHONE_FIELDS:
        phone = normalize_phone(record.data.get(field, ""))
        if phone:
            keys.append(f"PHONE:{phone}")
    one_id = normalize_id(record.data.get("ONE ID 序號", ""))
    if one_id:
        keys.append(f"ONEID:{one_id}")
    line_id = normalize_id(record.data.get("LINEID", ""))
    if line_id:
        keys.append(f"LINE:{line_id}")
    return list(dict.fromkeys(keys))


def merge_group(records: List[Record]) -> Tuple[Dict[str, str], Record]:
    ordered = sorted(records, key=lambda r: (r.priority, r.source_file.lower(), r.source_row))
    winner = ordered[0]
    merged = {col: clean_text(winner.data.get(col, "")) for col in CANONICAL_COLUMNS}
    for rec in ordered[1:]:
        for col in CANONICAL_COLUMNS:
            if not clean_text(merged.get(col, "")):
                merged[col] = clean_text(rec.data.get(col, ""))
    merged["來源優先權"] = str(winner.priority)
    merged["來源群組"] = winner.group_name
    merged["來源檔案"] = winner.source_file
    merged["來源列號"] = str(winner.source_row)
    merged["合併識別鍵"] = " | ".join(record_identifiers(winner))
    merged["更新時間"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return merged, winner


# ---------------------------------------------------------------------------
# Step 1 — merge_files
# g_files: List of (filename, file_bytes)
# existing_wb_bytes: bytes of existing workbook (or None for new file)
# ---------------------------------------------------------------------------

def merge_files(
    g_files: List[Tuple[str, bytes]],
    existing_wb_bytes: Optional[bytes] = None,
    filter_mode: str = "none",
) -> Tuple[bytes, Dict]:
    if not g_files:
        raise RuntimeError("未上傳任何 G1/G2/... 格式的 xlsx 檔案。")

    # Sort by priority then filename
    g_files_sorted = sorted(
        g_files,
        key=lambda t: (extract_priority(t[0]) or 9999, t[0].lower()),
    )

    records: List[Record] = []
    filtered_records: List[Record] = []
    manifest = []

    for filename, file_bytes in g_files_sorted:
        if extract_priority(filename) is None:
            continue
        file_records, file_filtered = read_excel_records_bytes(filename, file_bytes, filter_mode=filter_mode)
        records.extend(file_records)
        filtered_records.extend(file_filtered)
        manifest.append({
            "operation": "merge",
            "file_name": filename,
            "size": str(len(file_bytes)),
            "sha1": file_sha1_bytes(file_bytes),
            "processed_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    if not records:
        raise RuntimeError("有找到檔案，但讀不到可用資料列。請確認欄位名稱是否正確。")

    uf = UnionFind(len(records))
    key_owner: Dict[str, int] = {}
    for idx, record in enumerate(records):
        ids = record_identifiers(record)
        for key in ids:
            if key in key_owner:
                uf.union(idx, key_owner[key])
            else:
                key_owner[key] = idx

    groups: Dict[int, List[Record]] = defaultdict(list)
    for idx, record in enumerate(records):
        groups[uf.find(idx)].append(record)

    kept_rows: List[Dict[str, str]] = []
    dropped_rows: List[Dict[str, str]] = []
    for group_records in groups.values():
        merged_row, winner = merge_group(group_records)
        kept_rows.append(merged_row)
        for rec in group_records:
            if rec is winner:
                continue
            drop_row = {col: clean_text(rec.data.get(col, "")) for col in CANONICAL_COLUMNS}
            drop_row.update({
                "來源優先權": str(rec.priority),
                "來源群組": rec.group_name,
                "來源檔案": rec.source_file,
                "來源列號": str(rec.source_row),
                "合併識別鍵": " | ".join(record_identifiers(rec)),
                "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "保留來源": winner.source_file,
                "保留群組": winner.group_name,
            })
            dropped_rows.append(drop_row)

    kept_rows.sort(key=lambda r: (int(r.get("來源優先權") or 9999), r.get("來源檔案", ""), r.get("姓名", "")))

    if existing_wb_bytes:
        wb = _load_wb(existing_wb_bytes)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    headers = CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS
    write_dicts_sheet(wb, MERGED_SHEET, kept_rows, headers)
    write_dicts_sheet(wb, MASTER_SHEET, kept_rows, headers)
    write_dicts_sheet(wb, DUPLICATE_SHEET, dropped_rows, headers + ["保留來源", "保留群組"])
    filtered_rows = [
        {**{col: rec.data.get(col, "") for col in CANONICAL_COLUMNS},
         "來源優先權": str(rec.priority),
         "來源群組": rec.group_name,
         "來源檔案": rec.source_file,
         "來源列號": str(rec.source_row),
         "合併識別鍵": "",
         "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        for rec in filtered_records
    ]
    write_dicts_sheet(wb, FILTERED_DNC_SHEET, filtered_rows, headers)
    write_manifest_sheet(wb, manifest)
    write_readme_sheet(wb)
    if wb.sheetnames and wb.sheetnames[0] != README_SHEET:
        wb.move_sheet(wb[README_SHEET], offset=-(wb.index(wb[README_SHEET])))

    return _wb_to_bytes(wb), {
        "source_files": len(g_files_sorted),
        "input_rows": len(records),
        "filtered_rows": len(filtered_records),
        "kept_rows": len(kept_rows),
        "dropped_rows": len(dropped_rows),
    }


# ---------------------------------------------------------------------------
# Step 2 — remove_recent_orders
# order_files: List of (filename, file_bytes)  (.xlsx or .csv)
# ---------------------------------------------------------------------------

def _extract_recent_order_phones(
    order_files: List[Tuple[str, bytes]],
) -> Tuple[set, List[Dict[str, str]]]:
    phones: set = set()
    manifest = []
    for filename, file_bytes in order_files:
        manifest.append({
            "operation": "recent_orders",
            "file_name": filename,
            "size": str(len(file_bytes)),
            "sha1": file_sha1_bytes(file_bytes),
            "processed_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "xlsx":
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [clean_text(x) for x in rows[0]]
            phone_idx = None
            for idx, header in enumerate(headers):
                nh = normalize_header(header)
                if nh in {"MOBILE", "PHONE", "手機".upper(), "手機(CR)SMS".upper(), "手機(和泰會員)SMS".upper()}:
                    phone_idx = idx
                    break
            if phone_idx is None:
                continue
            for row in rows[1:]:
                value = row[phone_idx] if phone_idx < len(row) else ""
                phone = normalize_phone(value)
                if phone:
                    phones.add(phone)
        elif ext == "csv":
            _, csv_rows = read_csv_rows_bytes(file_bytes)
            phone_key = None
            for header in (csv_rows[0].keys() if csv_rows else []):
                nh = normalize_header(header)
                if nh in {"MOBILE", "PHONE", "手機".upper()}:
                    phone_key = header
                    break
            if not phone_key:
                continue
            for row in csv_rows:
                phone = normalize_phone(row.get(phone_key, ""))
                if phone:
                    phones.add(phone)
    return phones, manifest


def remove_recent_orders(
    wb_bytes: bytes,
    order_files: List[Tuple[str, bytes]],
) -> Tuple[bytes, Dict]:
    wb = _load_wb(wb_bytes)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    current_rows = sheet_to_dicts(wb[MASTER_SHEET])
    phones, manifest_rows = _extract_recent_order_phones(order_files)
    kept, removed = [], []
    for row in current_rows:
        matched = False
        matched_phone = ""
        for field in PHONE_FIELDS:
            phone = normalize_phone(row.get(field, ""))
            if phone and phone in phones:
                matched = True
                matched_phone = phone
                break
        if matched:
            log_row = dict(row)
            log_row["排除原因"] = "近年受訂"
            log_row["命中手機"] = matched_phone
            removed.append(log_row)
        else:
            kept.append(row)

    headers = list(dict.fromkeys(CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS + ["排除原因", "命中手機"]))
    write_dicts_sheet(wb, MASTER_SHEET, kept, CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS)
    write_dicts_sheet(wb, RECENT_REMOVED_SHEET, removed, headers)

    existing_manifest = sheet_to_dicts(wb[MANIFEST_SHEET]) if MANIFEST_SHEET in wb.sheetnames else []
    existing_manifest.extend(manifest_rows)
    write_manifest_sheet(wb, existing_manifest)

    return _wb_to_bytes(wb), {
        "before": len(current_rows),
        "removed": len(removed),
        "after": len(kept),
        "recent_order_phones": len(phones),
    }


# ---------------------------------------------------------------------------
# Step 3 — export_phone_template  →  returns CSV bytes
# ---------------------------------------------------------------------------

def export_phone_template(wb_bytes: bytes) -> Tuple[bytes, int]:
    wb = load_workbook(io.BytesIO(wb_bytes), read_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    rows = sheet_to_dicts(wb[MASTER_SHEET])
    seen: set = set()
    phones: List[str] = []
    for row in rows:
        for field in ["手機(CR)SMS", "手機(和泰會員)SMS"]:
            phone = normalize_phone(row.get(field, ""))
            if phone and phone not in seen:
                seen.add(phone)
                phones.append(phone)
                break

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Phone"])
    for phone in phones:
        writer.writerow([phone])

    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return csv_bytes, len(phones)


# ---------------------------------------------------------------------------
# Step 4 — collect_short_urls
# csv_bytes: the platform-returned CSV (No, Phone, Url, Count)
# ---------------------------------------------------------------------------

def collect_short_urls(
    wb_bytes: bytes,
    csv_bytes: bytes,
    csv_filename: str,
    column_name: str,
) -> Tuple[bytes, Dict]:
    wb = _load_wb(wb_bytes)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    _, csv_rows = read_csv_rows_bytes(csv_bytes)
    phone_to_url: Dict[str, Tuple[str, str]] = {}
    for row in csv_rows:
        phone = normalize_phone(row.get("Phone", ""))
        url = clean_text(row.get("Url", ""))
        count = clean_text(row.get("Count", ""))
        if phone and url and phone not in phone_to_url:
            phone_to_url[phone] = (url, count)

    count_column = f"{column_name}_次數"
    current_rows = sheet_to_dicts(wb[MASTER_SHEET])
    filled = 0
    unmatched = 0
    log_rows = []

    for row in current_rows:
        matched_url = ""
        matched_count = ""
        for field in ["手機(CR)SMS", "手機(和泰會員)SMS"]:
            phone = normalize_phone(row.get(field, ""))
            if phone and phone in phone_to_url:
                matched_url, matched_count = phone_to_url[phone]
                break

        if matched_url:
            if not row.get(column_name, ""):
                row[column_name] = matched_url
            row[count_column] = matched_count
            filled += 1
        else:
            row.setdefault(column_name, "")
            row.setdefault(count_column, "")
            unmatched += 1

        log_rows.append({
            "姓名": row.get("姓名", ""),
            "手機(CR)SMS": row.get("手機(CR)SMS", ""),
            "目標欄位": column_name,
            "短網址": row.get(column_name, ""),
            "次數": row.get(count_column, ""),
        })

    existing_keys = list(dict.fromkeys(k for row in current_rows for k in row.keys()))
    headers = list(dict.fromkeys(CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS + existing_keys + [column_name, count_column]))
    write_dicts_sheet(wb, MASTER_SHEET, current_rows, headers)

    log_headers = ["姓名", "手機(CR)SMS", "目標欄位", "短網址", "次數"]
    existing_log = sheet_to_dicts(wb[SHORT_URL_LOG_SHEET]) if SHORT_URL_LOG_SHEET in wb.sheetnames else []
    write_dicts_sheet(wb, SHORT_URL_LOG_SHEET, existing_log + log_rows, log_headers)

    manifest_row = {
        "operation": "collect_short_urls",
        "file_name": csv_filename,
        "size": str(len(csv_bytes)),
        "sha1": file_sha1_bytes(csv_bytes),
        "processed_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    existing_manifest = sheet_to_dicts(wb[MANIFEST_SHEET]) if MANIFEST_SHEET in wb.sheetnames else []
    existing_manifest.append(manifest_row)
    write_manifest_sheet(wb, existing_manifest)

    return _wb_to_bytes(wb), {"users": len(current_rows), "filled": filled, "unmatched": unmatched}


# ---------------------------------------------------------------------------
# Step 5 — match_stay_list
# xlsx_bytes: the 留名單 xlsx
# ---------------------------------------------------------------------------

def match_stay_list(
    wb_bytes: bytes,
    xlsx_bytes: bytes,
    xlsx_filename: str,
) -> Tuple[bytes, Dict]:
    wb_out = _load_wb(wb_bytes)
    if MASTER_SHEET not in wb_out.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    wb_src = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws_src = wb_src.active
    all_rows = list(ws_src.iter_rows(values_only=True))

    # Find header row: first row where ≥3 cells are non-empty
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if sum(1 for v in row if v not in (None, "")) >= 3:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError("留名單檔案找不到欄位標題列。")

    headers_src = [clean_text(v) for v in all_rows[header_row_idx]]
    phone_col = next((i for i, h in enumerate(headers_src) if normalize_header(h) == normalize_header("聯絡電話")), None)
    note_col = next((i for i, h in enumerate(headers_src) if normalize_header(h) == normalize_header("備註")), None)
    if phone_col is None or note_col is None:
        raise RuntimeError("留名單檔案找不到「聯絡電話」或「備註」欄位。")

    phone_set: set = set()
    lineid_set: set = set()
    for row in all_rows[header_row_idx + 1:]:
        note = clean_text(row[note_col] if note_col < len(row) else "")
        if "精準行銷" not in note:
            continue
        phone = normalize_phone(row[phone_col] if phone_col < len(row) else "")
        if phone:
            phone_set.add(phone)
        m = _LINE_ID_RE.search(note)
        if m:
            lineid_set.add(normalize_id(m.group(1)))

    current_rows = sheet_to_dicts(wb_out[MASTER_SHEET])
    sms_matched = 0
    line_matched = 0

    for row in current_rows:
        sms_hit = any(
            normalize_phone(row.get(f, "")) in phone_set
            for f in ["手機(CR)SMS", "手機(和泰會員)SMS"]
            if normalize_phone(row.get(f, ""))
        )
        row["簡訊留名單"] = "V" if sms_hit else "X"
        if sms_hit:
            sms_matched += 1

        line_id = normalize_id(row.get("LINEID", ""))
        line_hit = bool(line_id and line_id in lineid_set)
        row["LINE留名單"] = "V" if line_hit else "X"
        if line_hit:
            line_matched += 1

    headers = list(dict.fromkeys(CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS))
    write_dicts_sheet(wb_out, MASTER_SHEET, current_rows, headers)

    manifest_row = {
        "operation": "match_stay_list",
        "file_name": xlsx_filename,
        "size": str(len(xlsx_bytes)),
        "sha1": file_sha1_bytes(xlsx_bytes),
        "processed_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    existing_manifest = sheet_to_dicts(wb_out[MANIFEST_SHEET]) if MANIFEST_SHEET in wb_out.sheetnames else []
    existing_manifest.append(manifest_row)
    write_manifest_sheet(wb_out, existing_manifest)

    return _wb_to_bytes(wb_out), {"users": len(current_rows), "sms_matched": sms_matched, "line_matched": line_matched}
