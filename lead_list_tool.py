from __future__ import annotations

import csv
import hashlib
import os
import re
import sys
import traceback
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import Workbook, load_workbook

APP_VERSION = "0.1.0"

CANONICAL_COLUMNS = [
    "編號",
    "姓名",
    "ONE ID 序號",
    "LINEID",
    "手機(和泰會員)電銷",
    "手機(CR)電銷",
    "手機(和泰會員)SMS",
    "手機(CR)SMS",
    "訂單",
    "現保有車款_T",
    "現保有車交車年份_T",
    "現保有車款_L",
    "現保有車交車年份_L",
    "簡訊留名單",
    "LINE留名單",
]

OUTPUT_METADATA_COLUMNS = [
    "來源優先權",
    "來源群組",
    "來源檔案",
    "來源列號",
    "合併識別鍵",
    "更新時間",
]

MASTER_SHEET = "Working_List"
MERGED_SHEET = "Merged_Master"
DUPLICATE_SHEET = "Dropped_Duplicates"
FILTERED_DNC_SHEET = "Filtered_DNC"
RECENT_REMOVED_SHEET = "Removed_RecentOrders"
SHORT_URL_LOG_SHEET = "ShortURL_Log"
MANIFEST_SHEET = "Manifest"
README_SHEET = "README"

HEADER_ALIASES = {
    "編號": ["編號", "no", "序號"],
    "姓名": ["姓名", "name"],
    "ONE ID 序號": ["ONE ID 序號", "ONEID", "ONE ID", "ONE_ID", "ONE ID序號"],
    "LINEID": ["LINEID", "LINE ID", "LINE_ID"],
    "手機(和泰會員)電銷": ["手機(和泰會員)電銷", "手機(和泰會員) 電銷", "和泰會員電銷", "電話(和泰會員)電銷"],
    "手機(CR)電銷": ["手機(CR)電銷", "手機(CR) 電銷", "CR電銷", "電話(CR)電銷"],
    "手機(和泰會員)SMS": ["手機(和泰會員)SMS", "手機(和泰會員) SMS", "和泰會員SMS", "手機(和泰會員)簡訊"],
    "手機(CR)SMS": ["手機(CR)SMS", "手機(CR) SMS", "CRSMS", "手機(CR)簡訊", "手機(CR) sms"],
    "訂單": ["訂單", "MOBILE", "Phone", "PHONE", "手機", "行動電話"],
    "現保有車款_T": ["現保有車款_T", "現保有車款T"],
    "現保有車交車年份_T": ["現保有車交車年份_T", "現保有車交車年份T"],
    "現保有車款_L": ["現保有車款_L", "現保有車款L"],
    "現保有車交車年份_L": ["現保有車交車年份_L", "現保有車交車年份L"],
}

PHONE_FIELDS = [
    "手機(CR)SMS",
    "手機(和泰會員)SMS",
    "手機(CR)電銷",
    "手機(和泰會員)電銷",
]

PHONE_ELEC_FIELDS = ["手機(和泰會員)電銷", "手機(CR)電銷"]
PHONE_SMS_FIELDS = ["手機(和泰會員)SMS", "手機(CR)SMS"]

DNC_VALUES = {
    "不聯繫", "不連繫",
    "電話不聯繫", "電話不要打", "該電話不可聯繫",
    "簡訊不聯繫",
    "個資未授權",
}

NULL_LIKE = {
    "", "none", "null", "nan", "#n/a", "n/a",
    "電話不聯繫", "電話不要打", "該電話不可聯繫",
    "不聯繫", "不連繫",
    "簡訊不聯繫",
    "個資未授權",
    "無", "-", "--",
}

G_FILE_PATTERN = re.compile(r"(?:^|[^A-Z0-9])G(\d+)(?:_|\b|$)", re.IGNORECASE)


@dataclass
class Record:
    data: Dict[str, str]
    source_file: str
    source_path: str
    source_row: int
    priority: int
    group_name: str


class UnionFind:
    def __init__(self, n: int) -> None:
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x: int) -> int:
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a: int, b: int) -> None:
        ra = self.find(a)
        rb = self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            self.parent[ra] = rb
        elif self.rank[ra] > self.rank[rb]:
            self.parent[rb] = ra
        else:
            self.parent[rb] = ra
            self.rank[ra] += 1


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in NULL_LIKE:
        return ""
    return text



def normalize_header(text: str) -> str:
    text = clean_text(text)
    text = text.replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", "", text)
    return text.upper()



def normalize_phone(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    if len(digits) == 11 and digits.startswith("886"):
        digits = "0" + digits[-9:]
    elif len(digits) > 10 and digits.endswith(tuple(str(i) for i in range(10))):
        digits = digits[-10:]
    elif len(digits) == 9 and digits.startswith("9"):
        digits = "0" + digits
    if len(digits) < 8:
        return ""
    return digits



def normalize_id(value) -> str:
    text = clean_text(value)
    return text.upper()



def file_sha1(path: Path) -> str:
    h = hashlib.sha1()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()



def extract_priority(filename: str) -> Optional[int]:
    match = G_FILE_PATTERN.search(filename)
    if not match:
        return None
    return int(match.group(1))



def extract_group_name(filename: str) -> str:
    p = extract_priority(filename)
    if p is None:
        return ""
    return f"G{p}"



def map_headers(headers: List[str]) -> Dict[str, int]:
    normalized = [normalize_header(h) for h in headers]
    result: Dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        alias_set = {normalize_header(a) for a in aliases}
        for idx, header in enumerate(normalized):
            if header in alias_set:
                result[canonical] = idx
                break
    return result



def first_non_empty_row(rows: Iterable[Tuple]) -> Tuple[int, List[str], List[List[str]]]:
    cached_rows: List[List[str]] = []
    header_idx = 0
    header_row: List[str] = []
    for excel_row_idx, row in rows:
        values = [clean_text(cell) for cell in row]
        cached_rows.append(values)
        non_empty = sum(1 for v in values if v)
        if non_empty >= 2:
            header_idx = excel_row_idx
            header_row = values
            break
    return header_idx, header_row, cached_rows



def _is_dnc(raw_row: List[str], header_map: Dict[str, int], fields: List[str]) -> bool:
    for f in fields:
        if f not in header_map:
            continue
        idx = header_map[f]
        if idx < len(raw_row) and raw_row[idx].strip() in DNC_VALUES:
            return True
    return False


def read_excel_records(path: Path, filter_mode: str = "none") -> Tuple[List[Record], List[Record]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    row_iter = ((idx, list(row)) for idx, row in enumerate(ws.iter_rows(values_only=True), start=1))
    header_idx, header_row, _ = first_non_empty_row(row_iter)
    header_map = map_headers(header_row)
    if not header_map:
        return [], []

    results: List[Record] = []
    filtered: List[Record] = []
    priority = extract_priority(path.name) or 9999
    group_name = extract_group_name(path.name) or path.stem

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= header_idx:
            continue
        raw_row = [str(cell).strip() if cell is not None else "" for cell in row]
        row_values = [clean_text(cell) for cell in row]
        if not any(row_values):
            continue
        record_data = {col: "" for col in CANONICAL_COLUMNS}
        for canonical, col_idx in header_map.items():
            if col_idx < len(row_values):
                record_data[canonical] = clean_text(row_values[col_idx])
        if not any(record_data.values()):
            continue
        phones_empty = not any(record_data.get(f, "") for f in PHONE_FIELDS)
        line_empty = not record_data.get("LINEID", "")
        if phones_empty and line_empty:
            continue
        rec = Record(
            data=record_data,
            source_file=path.name,
            source_path=str(path),
            source_row=row_idx,
            priority=priority,
            group_name=group_name,
        )
        if filter_mode != "none":
            phone_dnc = _is_dnc(raw_row, header_map, PHONE_ELEC_FIELDS)
            sms_dnc = _is_dnc(raw_row, header_map, PHONE_SMS_FIELDS)
            should_filter = (
                (filter_mode == "phone" and phone_dnc)
                or (filter_mode == "sms" and sms_dnc)
                or (filter_mode == "either" and (phone_dnc or sms_dnc))
            )
            if should_filter:
                filtered.append(rec)
                continue
        results.append(rec)
    return results, filtered



def read_csv_rows(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    encodings = ["utf-8-sig", "utf-8", "cp950", "big5"]
    last_error = None
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                rows = [{k or "": clean_text(v) for k, v in row.items()} for row in reader]
                return headers, rows
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    raise RuntimeError(f"無法讀取 CSV：{path}\n{last_error}")



def record_identifiers(record: Record) -> List[str]:
    keys: List[str] = []
    for field in PHONE_FIELDS:
        phone = normalize_phone(record.data.get(field, ""))
        if phone:
            keys.append(f"PHONE:{phone}")
    one_id = normalize_id(record.data.get("ONE ID 序號", ""))
    if one_id:
        keys.append(f"ONEID:{one_id}")
    line_id = normalize_id(record.data.get("LINEID", ""))
    if line_id:
        keys.append(f"LINE:{line_id}")
    return list(dict.fromkeys(keys))



def best_record(records: List[Record]) -> Record:
    return sorted(records, key=lambda r: (r.priority, r.source_file.lower(), r.source_row))[0]



def merge_group(records: List[Record]) -> Tuple[Dict[str, str], Record]:
    ordered = sorted(records, key=lambda r: (r.priority, r.source_file.lower(), r.source_row))
    winner = ordered[0]
    merged = {col: clean_text(winner.data.get(col, "")) for col in CANONICAL_COLUMNS}
    for rec in ordered[1:]:
        for col in CANONICAL_COLUMNS:
            if not clean_text(merged.get(col, "")):
                merged[col] = clean_text(rec.data.get(col, ""))
    merged["來源優先權"] = str(winner.priority)
    merged["來源群組"] = winner.group_name
    merged["來源檔案"] = winner.source_file
    merged["來源列號"] = str(winner.source_row)
    merged["合併識別鍵"] = " | ".join(record_identifiers(winner))
    merged["更新時間"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return merged, winner



def auto_width(ws) -> None:
    widths: Dict[int, int] = defaultdict(lambda: 10)
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            text = clean_text(value)
            if not text:
                continue
            widths[idx] = min(max(widths[idx], len(text) + 2), 40)
    for idx, width in widths.items():
        ws.column_dimensions[_excel_col(idx)].width = width



def _excel_col(index: int) -> str:
    result = ""
    while index:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result



def recreate_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(title=name)



def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb



def write_readme_sheet(wb: Workbook) -> None:
    ws = recreate_sheet(wb, README_SHEET)
    rows = [
        ["本工具用途", "本機端整理名單，無須上傳外部服務"],
        ["版本", APP_VERSION],
        ["Working_List", "目前對外使用中的名單"],
        ["Merged_Master", "原始合併結果備份"],
        ["Dropped_Duplicates", "因優先權或重複被捨棄的名單"],
        ["Filtered_DNC", "因電話/SMS 標示不聯繫而被過濾的名單"],
        ["Removed_RecentOrders", "因近年受訂而被排除的名單"],
        ["ShortURL_Log", "短網址配對紀錄"],
        ["Manifest", "本次使用到的來源檔案與雜湊"],
    ]
    for row in rows:
        ws.append(row)
    auto_width(ws)



def write_manifest_sheet(wb: Workbook, manifest_rows: List[Dict[str, str]]) -> None:
    ws = recreate_sheet(wb, MANIFEST_SHEET)
    headers = ["operation", "file_name", "path", "size", "modified_time", "sha1"]
    ws.append(headers)
    for row in manifest_rows:
        ws.append([row.get(h, "") for h in headers])
    auto_width(ws)



def sheet_to_dicts(ws) -> List[Dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_text(h) for h in rows[0]]
    results = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else ""
            item[header] = clean_text(value)
        if any(item.values()):
            results.append(item)
    return results



def write_dicts_sheet(wb: Workbook, name: str, rows: List[Dict[str, str]], headers: List[str]) -> None:
    ws = recreate_sheet(wb, name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    auto_width(ws)
    ws.freeze_panes = "A2"



def scan_g_files(folder: Path) -> List[Path]:
    files = []
    for path in folder.iterdir():
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        if extract_priority(path.name) is None:
            continue
        files.append(path)
    return sorted(files, key=lambda p: (extract_priority(p.name) or 9999, p.name.lower()))



def merge_files(source_folder: Path, output_path: Path, filter_mode: str = "none") -> Dict[str, int]:
    g_files = scan_g_files(source_folder)
    if not g_files:
        raise RuntimeError("找不到任何 G1/G2/... 這類型的 xlsx 檔案。")

    records: List[Record] = []
    filtered_records: List[Record] = []
    manifest = []
    for file in g_files:
        file_records, file_filtered = read_excel_records(file, filter_mode=filter_mode)
        records.extend(file_records)
        filtered_records.extend(file_filtered)
        stat = file.stat()
        manifest.append(
            {
                "operation": "merge",
                "file_name": file.name,
                "path": str(file),
                "size": str(stat.st_size),
                "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "sha1": file_sha1(file),
            }
        )

    if not records:
        raise RuntimeError("有找到檔案，但讀不到可用資料列。請確認欄位名稱是否正確。")

    uf = UnionFind(len(records))
    key_owner: Dict[str, int] = {}
    for idx, record in enumerate(records):
        ids = record_identifiers(record)
        for key in ids:
            if key in key_owner:
                uf.union(idx, key_owner[key])
            else:
                key_owner[key] = idx

    groups: Dict[int, List[Record]] = defaultdict(list)
    for idx, record in enumerate(records):
        groups[uf.find(idx)].append(record)

    kept_rows: List[Dict[str, str]] = []
    dropped_rows: List[Dict[str, str]] = []
    for group_records in groups.values():
        merged_row, winner = merge_group(group_records)
        kept_rows.append(merged_row)
        for rec in group_records:
            if rec is winner:
                continue
            drop_row = {col: clean_text(rec.data.get(col, "")) for col in CANONICAL_COLUMNS}
            drop_row.update(
                {
                    "來源優先權": str(rec.priority),
                    "來源群組": rec.group_name,
                    "來源檔案": rec.source_file,
                    "來源列號": str(rec.source_row),
                    "合併識別鍵": " | ".join(record_identifiers(rec)),
                    "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "保留來源": winner.source_file,
                    "保留群組": winner.group_name,
                }
            )
            dropped_rows.append(drop_row)

    kept_rows.sort(key=lambda r: (int(r.get("來源優先權") or 9999), r.get("來源檔案", ""), r.get("姓名", "")))

    wb = ensure_workbook(output_path)
    headers = CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS
    write_dicts_sheet(wb, MERGED_SHEET, kept_rows, headers)
    write_dicts_sheet(wb, MASTER_SHEET, kept_rows, headers)
    write_dicts_sheet(wb, DUPLICATE_SHEET, dropped_rows, headers + ["保留來源", "保留群組"])
    filtered_rows = [
        {**{col: rec.data.get(col, "") for col in CANONICAL_COLUMNS},
         "來源優先權": str(rec.priority),
         "來源群組": rec.group_name,
         "來源檔案": rec.source_file,
         "來源列號": str(rec.source_row),
         "合併識別鍵": "",
         "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        for rec in filtered_records
    ]
    write_dicts_sheet(wb, FILTERED_DNC_SHEET, filtered_rows, headers)
    write_manifest_sheet(wb, manifest)
    write_readme_sheet(wb)
    if wb.sheetnames and wb.sheetnames[0] != README_SHEET:
        wb.move_sheet(wb[README_SHEET], offset=-(wb.index(wb[README_SHEET])))
    wb.save(output_path)
    return {
        "source_files": len(g_files),
        "input_rows": len(records),
        "filtered_rows": len(filtered_records),
        "kept_rows": len(kept_rows),
        "dropped_rows": len(dropped_rows),
    }



def extract_recent_order_phones(folder: Path) -> Tuple[set, List[Dict[str, str]]]:
    phones = set()
    manifest = []
    for path in folder.iterdir():
        if not path.is_file() or path.name.startswith("~$"):
            continue
        ext = path.suffix.lower()
        if ext not in {".xlsx", ".csv"}:
            continue
        stat = path.stat()
        manifest.append(
            {
                "operation": "recent_orders",
                "file_name": path.name,
                "path": str(path),
                "size": str(stat.st_size),
                "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "sha1": file_sha1(path),
            }
        )
        if ext == ".xlsx":
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [clean_text(x) for x in rows[0]]
            phone_idx = None
            for idx, header in enumerate(headers):
                nh = normalize_header(header)
                if nh in {"MOBILE", "PHONE", "手機".upper(), "手機(CR)SMS".upper(), "手機(和泰會員)SMS".upper()}:
                    phone_idx = idx
                    break
            if phone_idx is None:
                continue
            for row in rows[1:]:
                value = row[phone_idx] if phone_idx < len(row) else ""
                phone = normalize_phone(value)
                if phone:
                    phones.add(phone)
        else:
            headers, rows = read_csv_rows(path)
            phone_key = None
            for header in headers:
                nh = normalize_header(header)
                if nh in {"MOBILE", "PHONE", "手機".upper()}:
                    phone_key = header
                    break
            if not phone_key:
                continue
            for row in rows:
                phone = normalize_phone(row.get(phone_key, ""))
                if phone:
                    phones.add(phone)
    return phones, manifest



def remove_recent_orders(output_path: Path, recent_orders_folder: Path) -> Dict[str, int]:
    if not output_path.exists():
        raise RuntimeError("請先建立或選擇整理好的輸出檔案。")
    wb = load_workbook(output_path)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    current_rows = sheet_to_dicts(wb[MASTER_SHEET])
    phones, manifest_rows = extract_recent_order_phones(recent_orders_folder)
    kept, removed = [], []
    for row in current_rows:
        matched = False
        matched_phone = ""
        for field in PHONE_FIELDS:
            phone = normalize_phone(row.get(field, ""))
            if phone and phone in phones:
                matched = True
                matched_phone = phone
                break
        if matched:
            log_row = dict(row)
            log_row["排除原因"] = "近年受訂"
            log_row["命中手機"] = matched_phone
            removed.append(log_row)
        else:
            kept.append(row)

    headers = list(dict.fromkeys(CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS + ["排除原因", "命中手機"]))
    write_dicts_sheet(wb, MASTER_SHEET, kept, CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS)
    write_dicts_sheet(wb, RECENT_REMOVED_SHEET, removed, headers)

    existing_manifest = []
    if MANIFEST_SHEET in wb.sheetnames:
        existing_manifest = sheet_to_dicts(wb[MANIFEST_SHEET])
    existing_manifest.extend(manifest_rows)
    write_manifest_sheet(wb, existing_manifest)
    wb.save(output_path)
    return {
        "before": len(current_rows),
        "removed": len(removed),
        "after": len(kept),
        "recent_order_phones": len(phones),
    }



def export_phone_template(output_path: Path, template_output_path: Path) -> int:
    if not output_path.exists():
        raise RuntimeError("請先建立或選擇整理好的輸出檔案。")
    wb = load_workbook(output_path, read_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    rows = sheet_to_dicts(wb[MASTER_SHEET])
    seen: set = set()
    phones: List[str] = []
    for row in rows:
        for field in ["手機(CR)SMS", "手機(和泰會員)SMS"]:
            phone = normalize_phone(row.get(field, ""))
            if phone and phone not in seen:
                seen.add(phone)
                phones.append(phone)
                break

    with template_output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Phone"])
        for phone in phones:
            writer.writerow([phone])

    return len(phones)



def collect_short_urls(output_path: Path, source_csv_path: Path, column_name: str) -> Dict[str, int]:
    if not output_path.exists():
        raise RuntimeError("請先建立或選擇整理好的輸出檔案。")
    wb = load_workbook(output_path)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    _, csv_rows = read_csv_rows(source_csv_path)
    phone_to_url: Dict[str, Tuple[str, str]] = {}
    for row in csv_rows:
        phone = normalize_phone(row.get("Phone", ""))
        url = clean_text(row.get("Url", ""))
        count = clean_text(row.get("Count", ""))
        if phone and url and phone not in phone_to_url:
            phone_to_url[phone] = (url, count)

    count_column = f"{column_name}_次數"
    current_rows = sheet_to_dicts(wb[MASTER_SHEET])
    filled = 0
    unmatched = 0
    log_rows = []

    for row in current_rows:
        matched_url = ""
        matched_count = ""
        for field in ["手機(CR)SMS", "手機(和泰會員)SMS"]:
            phone = normalize_phone(row.get(field, ""))
            if phone and phone in phone_to_url:
                matched_url, matched_count = phone_to_url[phone]
                break

        if matched_url:
            if not row.get(column_name, ""):
                row[column_name] = matched_url
            row[count_column] = matched_count
            filled += 1
        else:
            row.setdefault(column_name, "")
            row.setdefault(count_column, "")
            unmatched += 1

        log_rows.append({
            "姓名": row.get("姓名", ""),
            "手機(CR)SMS": row.get("手機(CR)SMS", ""),
            "目標欄位": column_name,
            "短網址": row.get(column_name, ""),
            "次數": row.get(count_column, ""),
        })

    existing_keys = list(dict.fromkeys(k for row in current_rows for k in row.keys()))
    headers = list(dict.fromkeys(CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS + existing_keys + [column_name, count_column]))
    write_dicts_sheet(wb, MASTER_SHEET, current_rows, headers)

    log_headers = ["姓名", "手機(CR)SMS", "目標欄位", "短網址", "次數"]
    existing_log = sheet_to_dicts(wb[SHORT_URL_LOG_SHEET]) if SHORT_URL_LOG_SHEET in wb.sheetnames else []
    write_dicts_sheet(wb, SHORT_URL_LOG_SHEET, existing_log + log_rows, log_headers)

    stat = source_csv_path.stat()
    manifest_row = {
        "operation": "collect_short_urls",
        "file_name": source_csv_path.name,
        "path": str(source_csv_path),
        "size": str(stat.st_size),
        "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "sha1": file_sha1(source_csv_path),
    }
    existing_manifest = sheet_to_dicts(wb[MANIFEST_SHEET]) if MANIFEST_SHEET in wb.sheetnames else []
    existing_manifest.append(manifest_row)
    write_manifest_sheet(wb, existing_manifest)
    wb.save(output_path)
    return {"users": len(current_rows), "filled": filled, "unmatched": unmatched}


_LINE_ID_RE = re.compile(r"LINE\s*ID_?(U[A-Za-z0-9]+)", re.IGNORECASE)


def match_stay_list(output_path: Path, source_xlsx_path: Path) -> Dict[str, int]:
    if not output_path.exists():
        raise RuntimeError("請先建立或選擇整理好的輸出檔案。")
    wb_out = load_workbook(output_path)
    if MASTER_SHEET not in wb_out.sheetnames:
        raise RuntimeError("找不到 Working_List，請先執行『合併檔案』。")

    # Read step4 xlsx — headers are on row 2, skip row 1
    wb_src = load_workbook(source_xlsx_path, data_only=True, read_only=True)
    ws_src = wb_src.active
    all_rows = list(ws_src.iter_rows(values_only=True))

    # Find the header row: first row where at least 3 cells are non-empty
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if sum(1 for v in row if v not in (None, "")) >= 3:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError("留名單檔案找不到欄位標題列。")

    headers_src = [clean_text(v) for v in all_rows[header_row_idx]]
    phone_col = next((i for i, h in enumerate(headers_src) if normalize_header(h) == normalize_header("聯絡電話")), None)
    note_col = next((i for i, h in enumerate(headers_src) if normalize_header(h) == normalize_header("備註")), None)
    if phone_col is None or note_col is None:
        raise RuntimeError("留名單檔案找不到「聯絡電話」或「備註」欄位。")

    phone_set: set = set()
    lineid_set: set = set()
    for row in all_rows[header_row_idx + 1:]:
        note = clean_text(row[note_col] if note_col < len(row) else "")
        if "精準行銷" not in note:
            continue
        phone = normalize_phone(row[phone_col] if phone_col < len(row) else "")
        if phone:
            phone_set.add(phone)
        m = _LINE_ID_RE.search(note)
        if m:
            lineid_set.add(normalize_id(m.group(1)))

    current_rows = sheet_to_dicts(wb_out[MASTER_SHEET])
    sms_matched = 0
    line_matched = 0

    for row in current_rows:
        sms_hit = any(
            normalize_phone(row.get(f, "")) in phone_set
            for f in ["手機(CR)SMS", "手機(和泰會員)SMS"]
            if normalize_phone(row.get(f, ""))
        )
        row["簡訊留名單"] = "V" if sms_hit else "X"
        if sms_hit:
            sms_matched += 1

        line_id = normalize_id(row.get("LINEID", ""))
        line_hit = bool(line_id and line_id in lineid_set)
        row["LINE留名單"] = "V" if line_hit else "X"
        if line_hit:
            line_matched += 1

    headers = list(dict.fromkeys(CANONICAL_COLUMNS + OUTPUT_METADATA_COLUMNS))
    write_dicts_sheet(wb_out, MASTER_SHEET, current_rows, headers)

    stat = source_xlsx_path.stat()
    manifest_row = {
        "operation": "match_stay_list",
        "file_name": source_xlsx_path.name,
        "path": str(source_xlsx_path),
        "size": str(stat.st_size),
        "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "sha1": file_sha1(source_xlsx_path),
    }
    existing_manifest = sheet_to_dicts(wb_out[MANIFEST_SHEET]) if MANIFEST_SHEET in wb_out.sheetnames else []
    existing_manifest.append(manifest_row)
    write_manifest_sheet(wb_out, existing_manifest)
    wb_out.save(output_path)
    return {"users": len(current_rows), "sms_matched": sms_matched, "line_matched": line_matched}


class FilterModeDialog:
    """彈出視窗，讓使用者選擇合併時的不聯繫過濾模式。"""

    OPTIONS = [
        ("不過濾，全部匯入", "none"),
        ("電話（電銷）欄位標示不聯繫者，不匯入", "phone"),
        ("SMS 欄位標示不聯繫者，不匯入", "sms"),
        ("電話或 SMS 任一標示不聯繫者，不匯入（最嚴格）", "either"),
    ]

    def __init__(self, parent: tk.Tk) -> None:
        self.result: Optional[str] = None
        self._win = tk.Toplevel(parent)
        self._win.title("不聯繫過濾設定")
        self._win.resizable(False, False)
        self._win.grab_set()

        ttk.Label(self._win, text="合併時如何處理標示「不聯繫」的電話欄位？", padding=(16, 12)).pack(anchor="w")

        self._var = tk.StringVar(value="none")
        for label, value in self.OPTIONS:
            ttk.Radiobutton(self._win, text=label, variable=self._var, value=value).pack(anchor="w", padx=24, pady=2)

        btn_frame = ttk.Frame(self._win, padding=(16, 12))
        btn_frame.pack(fill="x")
        ttk.Button(btn_frame, text="確定", command=self._ok).pack(side="right", padx=(8, 0))
        ttk.Button(btn_frame, text="取消", command=self._cancel).pack(side="right")

        self._win.protocol("WM_DELETE_WINDOW", self._cancel)
        parent.wait_window(self._win)

    def _ok(self) -> None:
        self.result = self._var.get()
        self._win.destroy()

    def _cancel(self) -> None:
        self.result = None
        self._win.destroy()


class LeadListGUI:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("名單整理工具")
        self.root.geometry("840x560")
        self.mode = tk.StringVar(value="new")
        self.output_path_var = tk.StringVar()
        self._build()

    def _build(self) -> None:
        frame = ttk.Frame(self.root, padding=16)
        frame.pack(fill="both", expand=True)

        title = ttk.Label(frame, text="名單整理工具（mac / Windows）", font=("Arial", 16, "bold"))
        title.pack(anchor="w")

        note = ttk.Label(
            frame,
            text="本工具僅在本機端處理檔案，不會把資料上傳到外部服務。",
            foreground="#555555",
        )
        note.pack(anchor="w", pady=(4, 12))

        mode_frame = ttk.LabelFrame(frame, text="輸出模式", padding=10)
        mode_frame.pack(fill="x")
        ttk.Radiobutton(mode_frame, text="產生新檔案", variable=self.mode, value="new").grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(mode_frame, text="更新既有檔案", variable=self.mode, value="update").grid(row=0, column=1, sticky="w")

        output_frame = ttk.LabelFrame(frame, text="輸出檔案", padding=10)
        output_frame.pack(fill="x", pady=(12, 0))
        ttk.Entry(output_frame, textvariable=self.output_path_var).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        output_frame.columnconfigure(0, weight=1)
        ttk.Button(output_frame, text="選擇檔案", command=self.choose_output_path).grid(row=0, column=1)

        action_frame = ttk.LabelFrame(frame, text="功能", padding=10)
        action_frame.pack(fill="x", pady=(12, 0))
        ttk.Button(action_frame, text="1. 合併檔案", command=self.do_merge).grid(row=0, column=0, padx=(0, 8), pady=4, sticky="ew")
        ttk.Button(action_frame, text="2. 去除近年訂過車名單", command=self.do_remove_recent).grid(row=0, column=1, padx=(0, 8), pady=4, sticky="ew")
        ttk.Button(action_frame, text="3. 產生手機號碼 Template", command=self.do_export_template).grid(row=0, column=2, padx=(0, 8), pady=4, sticky="ew")
        ttk.Button(action_frame, text="4. 匯入短網址結果", command=self.do_collect_urls).grid(row=0, column=3, padx=(0, 8), pady=4, sticky="ew")
        ttk.Button(action_frame, text="5. 比對留名單", command=self.do_match_stay_list).grid(row=0, column=4, pady=4, sticky="ew")

        for idx in range(5):
            action_frame.columnconfigure(idx, weight=1)

        log_frame = ttk.LabelFrame(frame, text="執行紀錄", padding=10)
        log_frame.pack(fill="both", expand=True, pady=(12, 0))
        self.log_text = tk.Text(log_frame, wrap="word")
        self.log_text.pack(fill="both", expand=True)

    def log(self, text: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {text}\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def choose_output_path(self) -> None:
        if self.mode.get() == "new":
            path = filedialog.asksaveasfilename(
                title="選擇輸出檔案位置",
                defaultextension=".xlsx",
                filetypes=[("Excel 檔案", "*.xlsx")],
            )
        else:
            path = filedialog.askopenfilename(
                title="選擇要更新的檔案",
                filetypes=[("Excel 檔案", "*.xlsx")],
            )
        if path:
            self.output_path_var.set(path)
            self.log(f"已選擇輸出檔案：{path}")

    def get_output_path(self) -> Path:
        path_text = self.output_path_var.get().strip()
        if not path_text:
            self.choose_output_path()
            path_text = self.output_path_var.get().strip()
        if not path_text:
            raise RuntimeError("尚未選擇輸出檔案。")
        path = Path(path_text)
        if self.mode.get() == "update" and not path.exists():
            raise RuntimeError("更新模式下，請選擇既有的 xlsx 檔案。")
        return path

    def run_action(self, func) -> None:
        try:
            func()
        except Exception as exc:  # noqa: BLE001
            self.log(f"失敗：{exc}")
            messagebox.showerror("執行失敗", f"{exc}\n\n{traceback.format_exc()}")

    def do_merge(self) -> None:
        def _run():
            output_path = self.get_output_path()
            dlg = FilterModeDialog(self.root)
            if dlg.result is None:
                return
            filter_mode = dlg.result
            filter_labels = {
                "none": "不過濾",
                "phone": "過濾電話不聯繫",
                "sms": "過濾 SMS 不聯繫",
                "either": "電話或 SMS 任一不聯繫即過濾",
            }
            folder = filedialog.askdirectory(title="選擇 G1/G2... 檔案所在資料夾")
            if not folder:
                return
            self.log(f"開始合併資料夾：{folder}（過濾模式：{filter_labels.get(filter_mode, filter_mode)}）")
            result = merge_files(Path(folder), output_path, filter_mode=filter_mode)
            filtered_note = f"、過濾不聯繫 {result['filtered_rows']} 筆" if result['filtered_rows'] else ""
            self.log(
                f"完成。來源檔案 {result['source_files']} 份，"
                f"讀入 {result['input_rows'] + result['filtered_rows']} 筆"
                f"{filtered_note}，"
                f"輸出 {result['kept_rows']} 筆（去除重複 {result['dropped_rows']} 筆）。"
            )
            messagebox.showinfo("完成", f"合併完成\n輸出：{output_path}")
        self.run_action(_run)

    def do_remove_recent(self) -> None:
        def _run():
            output_path = self.get_output_path()
            folder = filedialog.askdirectory(title="選擇近年受訂名單所在資料夾")
            if not folder:
                return
            self.log(f"開始去除近年受訂名單：{folder}")
            result = remove_recent_orders(output_path, Path(folder))
            self.log(f"完成。原本 {result['before']} 筆，移除 {result['removed']} 筆，剩餘 {result['after']} 筆。")
            messagebox.showinfo("完成", f"去除近年受訂完成\n輸出：{output_path}")
        self.run_action(_run)

    def do_export_template(self) -> None:
        def _run():
            output_path = self.get_output_path()
            save_path = filedialog.asksaveasfilename(
                title="儲存手機號碼 Template",
                defaultextension=".csv",
                filetypes=[("CSV 檔案", "*.csv")],
                initialfile="手機號碼_template.csv",
            )
            if not save_path:
                return
            self.log(f"開始產生手機號碼 Template...")
            count = export_phone_template(output_path, Path(save_path))
            self.log(f"完成。共匯出 {count} 筆手機號碼。")
            messagebox.showinfo("完成", f"Template 已儲存：{save_path}")
        self.run_action(_run)

    def do_match_stay_list(self) -> None:
        def _run():
            output_path = self.get_output_path()
            xlsx_path = filedialog.askopenfilename(
                title="選擇留名單 xlsx",
                filetypes=[("Excel 檔案", "*.xlsx")],
            )
            if not xlsx_path:
                return
            self.log(f"開始比對留名單：{xlsx_path}")
            result = match_stay_list(output_path, Path(xlsx_path))
            self.log(f"完成。共 {result['users']} 筆，簡訊留名單命中 {result['sms_matched']} 筆，LINE 留名單命中 {result['line_matched']} 筆。")
            messagebox.showinfo("完成", f"留名單比對完成\n輸出：{output_path}")
        self.run_action(_run)

    def do_collect_urls(self) -> None:
        def _run():
            output_path = self.get_output_path()
            column_name = simpledialog.askstring("短網址欄位", "請輸入欄位名稱（將產生「名稱」和「名稱_次數」兩欄）：", parent=self.root)
            if not column_name:
                return
            csv_path = filedialog.askopenfilename(
                title="選擇平台回傳的短網址 CSV",
                filetypes=[("CSV 檔案", "*.csv")],
            )
            if not csv_path:
                return
            self.log(f"開始匯入短網址，欄位：{column_name}，來源：{csv_path}")
            result = collect_short_urls(output_path, Path(csv_path), column_name.strip())
            self.log(f"完成。共 {result['users']} 筆，命中 {result['filled']} 筆，未命中 {result['unmatched']} 筆。")
            messagebox.showinfo("完成", f"短網址匯入完成\n輸出：{output_path}")
        self.run_action(_run)



def run_app() -> None:
    root = tk.Tk()
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:  # noqa: BLE001
        pass
    app = LeadListGUI(root)
    app.log("程式已啟動。請先選擇『產生新檔案』或『更新既有檔案』。")
    root.mainloop()


if __name__ == "__main__":
    run_app()
